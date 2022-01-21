#######################################################
# QRコード作成
#######################################################
import qrcode
import openpyxl
import sys
from PIL import Image, ImageDraw, ImageFont
import shutil
import os

# excelファイルを受取、観客リストを返す関数


def input_data(exel_file_name):
    wb = openpyxl.load_workbook(exel_file_name, data_only=True)
    sheetnames_list = wb.sheetnames
    for sheetname in sheetnames_list:
        sheet = wb.get_sheet_by_name(str(sheetname))
        audience_data = []
        folder_data_input = []
        for row in range(2, sheet.max_row+1):
            data_input = {}
            data_input["name"] = sheet[str("a") + str(row)].value
            data_input["group"] = sheet[str("b") + str(row)].value
            data_input["token"] = sheet[str("c") + str(row)].value
            data_input["seat_num"] = sheet[str("d") + str(row)].value
            print(data_input)
            audience_data.append(data_input)
            folder_data_input.append(data_input["group"])
        folder_data = set(folder_data_input)
        data_list = {"audience_data": audience_data,
                     "folder_data": folder_data}
    print(data_list)
    return data_list


# 画像と観客データを受け取ってQRコード入りの画像を返す関数
def write_qr(image, token):
    qr = qrcode.QRCode(
        version=4,
        box_size=47,
        border=8
    )
    qr.add_data(token)
    qr.make()
    image_qr = qr.make_image(fill_color="black", back_color="white").convert('L')
    W_im, H_im = image.size
    W_qr, H_qr = image_qr.size
    qr_position_w = (W_im - W_qr)/2
    image.paste(image_qr,(int(qr_position_w), 1500))
    return image

# 画像を受取り、氏名シート名を記入して返す関数
def write_name(image, name, seat_num):
    draw = ImageDraw.Draw(image)
    font_path = "/usr/share/fonts/ヒラギノ角ゴシック W7.ttc"
    font_size = 1000
    W_im, H_im = image.size
    font = ImageFont.truetype(font_path, int(font_size))
    seat_font = ImageFont.truetype(font_path, int(230))
    w_text, h_text = draw.textsize(str(name), font)
    
    while w_text > W_im*2/3:
        font_size = font_size - 5
        font = ImageFont.truetype(font_path, int(font_size))
        w_text, h_text = draw.textsize(str(name), font)

    draw.text((W_im/2, 4500), "座席番号：" + str(seat_num), fill="black", font = seat_font, anchor='mb')
    draw.text((W_im/2, 5000), str(name), fill="black", font = font, anchor='mb')
    return image

# 画像と観客データを受取グループごとに保存する関数
def save_image(image, group, name):
    image.save("./output/" + group + "/" + group + "_" + name + ".png")


def mkdir_group(folder_list):
    shutil.rmtree("./output")
    os.mkdir("./output")
    for group in folder_list:
        os.mkdir('./output/' + str(group))



def main():
    image_file_name = sys.argv[1]
    print(image_file_name)
    exel_file_name = sys.argv[2]
    print(exel_file_name)

    # excelファイルから観客リストを取得
    input_data_list = input_data(exel_file_name)
    audience_list = input_data_list["audience_data"]
    group_list = input_data_list["folder_data"]

    #団体ごとのフォルダ作成
    mkdir_group(group_list)


    # 観客数でループ
    for audience_data in audience_list:

        # 書き込み様のデータをセット
        im = Image.open(image_file_name)

        # QRコードを画像に書き込み
        wroute_qr = write_qr(im, audience_data["token"])

        # 氏名を画像に書き込み
        wrote_qr_name = write_name(wroute_qr, audience_data["name"], audience_data["seat_num"])
        print(audience_data["name"], "を作成しました。")

        # 画像を保存
        save_image(wrote_qr_name, audience_data["group"], audience_data["name"])


if __name__ == '__main__':
    main()
    #im = Image.open("./dera_ticket.png")
    #audience_data = {"name": "佐久間", "group": "HOT", "token": "jump-jump.org"}
    #image = write_name(im, audience_data["name"])
    #image.save("./test.png")
